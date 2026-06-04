"""
BuyICT email scraper
Scrapes all contact email addresses from www.buyict.gov.au opportunities,
maintains a persistent Excel registry of known emails, emails a daily
report, and generates an index.html for GitHub Pages.

Config: set GMAIL_APP_PASSWORD environment variable, or edit CONFIG below.
"""

import asyncio
import html as html_mod
import json
import re
import smtplib
import logging
import sys
from datetime import datetime, timezone, timedelta

# Always use AEST (UTC+10) for display dates so GitHub Actions runner
# shows the correct Australian date regardless of server timezone.
AEST = timezone(timedelta(hours=10))
def now_aest(): return datetime.now(AEST)
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, Page

# ── Configuration ──────────────────────────────────────────────────────────────
CONFIG = {
    "from_email": "nick.claude.agents@gmail.com",
    "to_email": "nick.chapman@parbery.com.au",
    "gmail_app_password": "",  # Set here OR via GMAIL_APP_PASSWORD env var
    "base_url": "https://www.buyict.gov.au",
    "listing_url": "https://www.buyict.gov.au/public?id=opportunities",
    "rectangle_id": "bc93ec0a8738d950f973a8e50cbb3598",
    "concurrency": 8,
    "page_load_timeout": 30000,
    "log_file": str(Path(__file__).parent / "buyict_scraper.log"),
    "registry_file": str(Path(__file__).parent / "buyict_email_registry.xlsx"),
    "html_file": str(Path(__file__).parent / "index.html"),
    "pages_url": "https://nick-claude-agents.github.io/au-gov-buyict/",
}

EMAIL_RE = re.compile(r'\b[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}\b')
EXCLUDE_DOMAINS = {"buyict.gov.au", "servicenow.com", "example.com"}

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(CONFIG["log_file"], encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ── Excel registry ────────────────────────────────────────────────────────────

HEADERS   = ["Email Address", "Buyer", "Date Added"]
COL_WIDTHS = [34, 50, 14]

# Palette — aligned with au-gov-tenders
FILL_HEADER = PatternFill("solid", fgColor="19473C")   # dark green
FILL_NEW    = PatternFill("solid", fgColor="FFF3F4")   # pink tint
FILL_KNOWN  = PatternFill("solid", fgColor="FFFFFF")
FONT_HEADER = Font(color="FFFFFF", bold=True)
FONT_NEW    = Font(bold=True, color="C0004A")           # deep pink
THIN_BORDER = Border(
    left=Side(style="thin", color="D0E0DC"),
    right=Side(style="thin", color="D0E0DC"),
    top=Side(style="thin", color="D0E0DC"),
    bottom=Side(style="thin", color="D0E0DC"),
)


def _load_registry() -> dict[str, dict]:
    path = Path(CONFIG["registry_file"])
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    known: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            known[str(row[0]).lower()] = {
                "email": row[0],
                "buyer": row[1],
                "date_added": row[2],
            }
    return known


def update_registry(results: list[dict]) -> tuple[set[str], set[str]]:
    """Update the Excel registry. Returns (new_emails, known_emails)."""
    existing = _load_registry()
    today = now_aest().strftime("%Y-%m-%d")
    new_emails: set[str] = set()

    for result in results:
        for email in result["emails"]:
            el = email.lower()
            if el not in existing:
                new_emails.add(el)
                existing[el] = {
                    "email": email,
                    "buyer": result["buyer"],
                    "date_added": today,
                }

    all_rows = sorted(
        existing.values(),
        key=lambda r: (r.get("date_added", ""), r.get("email", "")),
        reverse=True,
    )

    wb = openpyxl.Workbook()
    ws = wb.active

    for col, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for row_idx, r in enumerate(all_rows, 2):
        is_new = r["email"].lower() in new_emails
        values = [r.get("email", ""), r.get("buyer", ""), r.get("date_added", "")]
        fill = FILL_NEW if is_new else FILL_KNOWN
        font = FONT_NEW if is_new else Font()
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.font = font
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row_idx].height = 16

    ws.title = f"Registry ({len(all_rows)} emails)"
    wb.save(CONFIG["registry_file"])
    log.info(f"Registry updated: {len(all_rows)} total, {len(new_emails)} new -> {CONFIG['registry_file']}")
    return new_emails, set(existing.keys()) - new_emails


# ── Step 1: Collect all opportunity URLs ──────────────────────────────────────

async def collect_all_opportunity_urls(page: Page) -> list[dict]:
    items = []
    responses: list[str] = []

    async def on_response(resp):
        if f"rectangle/{CONFIG['rectangle_id']}" in resp.url:
            try:
                body = await resp.text()
                responses.append(body)
            except Exception:
                pass

    page.on("response", on_response)

    log.info("Loading opportunities listing page...")
    await page.goto(CONFIG["listing_url"], wait_until="networkidle",
                    timeout=CONFIG["page_load_timeout"])

    initial = json.loads(responses[0])
    wd = initial["result"]["data"]
    total = int(wd["totalItems"])
    size = int(wd["page_size"])
    total_pages = (total + size - 1) // size
    log.info(f"Found {total} opportunities across {total_pages} pages")

    items.extend(wd["pageItems"])

    seen_ids = {item["sys_id"] for item in items}

    for pg in range(2, total_pages + 1):
        page_items = []
        # Retry up to 3 times if we get all-duplicate data (stale Angular response)
        for attempt in range(3):
            responses.clear()
            await page.evaluate(f"""
                () => {{
                    const pag = document.querySelector('.pagination');
                    const scope = angular.element(pag).scope();
                    scope.selectPage({pg}, null);
                    scope.$apply();
                }}
            """)
            # Wait up to 6s for a response
            for _ in range(30):
                await asyncio.sleep(0.2)
                if responses:
                    break

            if not responses:
                log.warning(f"  Page {pg}: no response (attempt {attempt+1})")
                await asyncio.sleep(1)
                continue

            data = json.loads(responses[0])
            page_items = data["result"]["data"].get("pageItems", [])
            new_ids = {item["sys_id"] for item in page_items if item["sys_id"] not in seen_ids}

            if new_ids or not page_items:
                break  # Got fresh data (or empty page at end)
            else:
                log.warning(f"  Page {pg}: all {len(page_items)} items are duplicates "
                            f"(attempt {attempt+1}) — retrying...")
                await asyncio.sleep(1)

        new_count = sum(1 for item in page_items if item["sys_id"] not in seen_ids)
        items.extend(page_items)
        seen_ids.update(item["sys_id"] for item in page_items)
        log.info(f"  Page {pg}/{total_pages}: {len(page_items)} items "
                 f"({new_count} new, total so far: {len(items)})")

    # Deduplicate by sys_id — pagination occasionally returns the same item twice
    seen = set()
    unique_items = []
    for item in items:
        sid = item.get("sys_id")
        if sid and sid not in seen:
            seen.add(sid)
            unique_items.append(item)
    if len(unique_items) < len(items):
        log.info(f"Deduplicated: {len(items)} -> {len(unique_items)} items ({len(items)-len(unique_items)} dupes removed)")
    log.info(f"Collected {len(unique_items)} opportunity URLs")
    return unique_items


# ── Step 2: Extract emails from a detail page ─────────────────────────────────

async def extract_emails_from_detail(context, item: dict) -> dict:
    url = CONFIG["base_url"] + "/public" + item["url"]
    number = item.get("number", "?")
    title = item.get("short_description", "")[:80]

    page = await context.new_page()
    emails = set()
    try:
        await page.goto(url, wait_until="networkidle",
                        timeout=CONFIG["page_load_timeout"])
        content = await page.content()
        found = EMAIL_RE.findall(content)
        emails = {
            e.lower() for e in found
            if e.split("@")[-1].lower() not in EXCLUDE_DOMAINS
        }
    except Exception as e:
        log.warning(f"  {number}: failed to load ({e})")
    finally:
        await page.close()

    if emails:
        log.info(f"  {number}: {emails}")
    return {
        "number": number,
        "title": title,
        "url": url,
        "buyer": item.get("u_buyer_name", ""),
        "close_date": item.get("u_close_date", ""),
        "emails": sorted(emails),
    }


# ── Step 3: Run all detail pages concurrently ─────────────────────────────────

async def scrape_all_details(context, items: list[dict]) -> list[dict]:
    sem = asyncio.Semaphore(CONFIG["concurrency"])
    results = []

    async def bounded(item):
        async with sem:
            return await extract_emails_from_detail(context, item)

    tasks = [asyncio.create_task(bounded(item)) for item in items]
    for i, coro in enumerate(asyncio.as_completed(tasks), 1):
        result = await coro
        results.append(result)
        if i % 10 == 0:
            log.info(f"Progress: {i}/{len(items)} detail pages done")

    return results


# ── Step 4: Generate HTML page ────────────────────────────────────────────────

def generate_html(results: list[dict], new_emails: set[str]) -> str:
    run_dt  = now_aest()
    run_date = run_dt.strftime("%d %B %Y")
    run_ts   = run_dt.strftime("%Y-%m-%d %H:%M")
    with_emails    = [r for r in results if r["emails"]]
    without_emails = [r for r in results if not r["emails"]]
    all_emails_set = {e for r in with_emails for e in r["emails"]}

    def esc(s): return html_mod.escape(str(s))

    def sort_key(r):
        try:
            return datetime.strptime(r["close_date"], "%d-%m-%Y")
        except Exception:
            return datetime.max

    rows = ""
    for r in sorted(results, key=sort_key):
        try:
            sort_date = datetime.strptime(r["close_date"], "%d-%m-%Y").strftime("%Y-%m-%d")
            disp_date = datetime.strptime(r["close_date"], "%d-%m-%Y").strftime("%d %b %Y")
        except Exception:
            sort_date = r["close_date"]
            disp_date = r["close_date"]
        if r["emails"]:
            email_cells = []
            for email in r["emails"]:
                if email.lower() in new_emails:
                    email_cells.append(
                        f'<a href="mailto:{esc(email)}" class="email-new" title="New this run">'
                        f'&#9733;&nbsp;{esc(email)}</a>'
                    )
                else:
                    email_cells.append(
                        f'<a href="mailto:{esc(email)}" class="email-link">{esc(email)}</a>'
                    )
            emails_html = "<br>".join(email_cells)
        else:
            emails_html = ""
        is_new = any(e.lower() in new_emails for e in r["emails"])
        new_row = ' class="row-new"' if is_new else ""
        rows += (
            f'<tr{new_row}>'
            f'<td><a href="{esc(r["url"])}" target="_blank" rel="noopener">{esc(r["number"])}</a></td>'
            f'<td>{esc(r["buyer"])}</td>'
            f'<td>{esc(r["title"])}</td>'
            f'<td data-sort="{sort_date}">{disp_date}</td>'
            f'<td>{emails_html}</td>'
            f'</tr>\n'
        )

    no_email_rows = ""  # all rows now rendered in the single loop above

    new_banner = ""
    if new_emails:
        n = len(new_emails)
        email_list = ", ".join(
            f'<a href="mailto:{e}">{e}</a>' for e in sorted(new_emails)
        )
        new_banner = (
            f'<div class="banner-new">&#9733; '
            f'<strong>{n} new contact{"s" if n != 1 else ""} this run:</strong> {email_list}</div>'
        )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>BuyICT Opportunities — {run_date}</title>
<style>
  :root {{
    --dark:       #19473c;
    --mid:        #1f735e;
    --pink:       #ffadb5;
    --pink-bg:    #fff3f4;
    --border:     #d0e0dc;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, sans-serif; font-size: 14px; color: #222; background: #f0f5f4; }}
  header {{ background: var(--dark); color: #fff; padding: 18px 24px; }}
  header h1 {{ font-size: 20px; font-weight: bold; }}
  header p {{ font-size: 13px; color: #a8cec7; margin-top: 4px; }}
  header a {{ color: #a8cec7; }}
  .container {{ max-width: 1400px; margin: 0 auto; padding: 20px 16px; }}
  .stats {{ display: flex; gap: 16px; flex-wrap: wrap; margin-bottom: 16px; }}
  .stat {{ background: #fff; border: 1px solid var(--border); border-radius: 6px;
           padding: 10px 18px; text-align: center; }}
  .stat strong {{ display: block; font-size: 22px; color: var(--dark); }}
  .stat span {{ font-size: 12px; color: #888; }}
  .banner-new {{ background: var(--pink-bg); border-left: 4px solid var(--pink);
                 padding: 10px 14px; margin-bottom: 16px; border-radius: 0 6px 6px 0;
                 font-size: 13px; }}
  .search-bar {{ margin-bottom: 12px; }}
  .search-bar input {{ width: 100%; max-width: 400px; padding: 7px 12px;
                       border: 1px solid var(--border); border-radius: 4px; font-size: 14px; }}
  .search-bar input:focus {{ outline: none; border-color: var(--mid);
                              box-shadow: 0 0 0 2px rgba(31,115,94,0.15); }}
  .table-wrap {{ overflow-x: auto; background: #fff; border-radius: 6px;
                 border: 1px solid var(--border); }}
  table {{ width: 100%; border-collapse: collapse; }}
  thead tr {{ background: var(--dark); color: #fff; }}
  th {{ padding: 9px 10px; text-align: left; font-size: 13px; cursor: pointer; white-space: nowrap; }}
  th:hover {{ background: var(--mid); }}
  td {{ padding: 7px 10px; border-bottom: 1px solid #e8f0ee; vertical-align: top; }}
  td:nth-child(4) {{ white-space: nowrap; }}
  tr:last-child td {{ border-bottom: none; }}
  tr.row-new {{ background: var(--pink-bg); }}
  tr.row-no-email td {{ color: #aaa; }}
  tr:hover td {{ background: #eef5f3; }}
  tr.row-new:hover td {{ background: #ffe8ea; }}
  a {{ color: var(--mid); text-decoration: none; }}
  a:hover {{ text-decoration: underline; }}
  .email-new {{ color: #c0004a; font-weight: bold; }}
  .footer {{ text-align: center; color: #aaa; font-size: 12px; margin-top: 24px; padding-bottom: 24px; }}
  .hidden {{ display: none; }}
  th .sort-icon {{ font-size: 10px; margin-left: 4px; opacity: 0.6; }}
  .new-badge {{ background: var(--pink); color: #8b0030; font-size: 11px; font-weight: bold;
                padding: 1px 6px; border-radius: 3px; margin-left: 6px; vertical-align: middle; }}
</style>
</head>
<body>
<header>
  <h1>BuyICT — Current Opportunity Contacts</h1>
  <p>Scraped from <a href="https://www.buyict.gov.au/public?id=opportunities" target="_blank">buyict.gov.au</a>
     &nbsp;|&nbsp; Updated {run_ts} AEST &nbsp;|&nbsp;
     <a href="{CONFIG['pages_url']}">Permalink</a></p>
</header>

<div class="container">
  <div class="stats">
    <div class="stat"><strong>{len(results)}</strong><span>opportunities scraped</span></div>
    <div class="stat"><strong>{len(with_emails)}</strong><span>with contact email</span></div>
    <div class="stat"><strong>{len(all_emails_set)}</strong><span>unique addresses</span></div>
    <div class="stat"><strong>{len(new_emails) if new_emails else "&#8212;"}</strong><span>new contacts</span></div>
  </div>

  {new_banner}

  <div class="search-bar">
    <input type="text" id="searchInput" placeholder="Filter by buyer, title, email&hellip;" oninput="filterTable()">
  </div>

  <div class="table-wrap">
    <table id="mainTable">
      <thead>
        <tr>
          <th onclick="sortTable(0)">ID <span class="sort-icon">&#8645;</span></th>
          <th onclick="sortTable(1)">Buyer <span class="sort-icon">&#8645;</span></th>
          <th onclick="sortTable(2)">Title <span class="sort-icon">&#8645;</span></th>
          <th onclick="sortTable(3)">Closes <span class="sort-icon">&#8645;</span></th>
          <th onclick="sortTable(4)">Contact Email <span class="sort-icon">&#8645;</span></th>
        </tr>
      </thead>
      <tbody id="tableBody">
        {rows}
      </tbody>
    </table>
  </div>

  <div class="footer">
    &#9733; = new contact not seen in previous run &nbsp;|&nbsp;
    Data sourced from <a href="https://www.buyict.gov.au">buyict.gov.au</a> &nbsp;|&nbsp;
    Generated {run_ts}
  </div>
</div>

<script>
function filterTable() {{
  const q = document.getElementById('searchInput').value.toLowerCase();
  document.querySelectorAll('#tableBody tr').forEach(row => {{
    row.classList.toggle('hidden', q && !row.textContent.toLowerCase().includes(q));
  }});
}}
let sortDir = {{}};
function sortTable(col) {{
  const tbody = document.getElementById('tableBody');
  const rows = Array.from(tbody.querySelectorAll('tr'));
  const asc = sortDir[col] !== true;
  sortDir = {{}};
  sortDir[col] = asc;
  document.querySelectorAll('th .sort-icon').forEach((el, i) => {{
    el.textContent = i === col ? (asc ? '↑' : '↓') : '⇅';
  }});
  rows.sort((a, b) => {{
    const ac = a.cells[col], bc = b.cells[col];
    const av = ac?.dataset.sort ?? ac?.textContent.trim() ?? '';
    const bv = bc?.dataset.sort ?? bc?.textContent.trim() ?? '';
    return asc ? av.localeCompare(bv) : bv.localeCompare(av);
  }});
  rows.forEach(r => tbody.appendChild(r));
}}
window.addEventListener('DOMContentLoaded', () => sortTable(3));
</script>
</body>
</html>
"""


def save_html(html: str) -> None:
    path = Path(CONFIG["html_file"])
    path.write_text(html, encoding="utf-8")
    log.info(f"HTML written to {path}")


# ── Step 5: Send email ────────────────────────────────────────────────────────

def send_email(results: list[dict], new_emails: set[str], password: str):
    run_date = now_aest().strftime("%d %B %Y")
    with_emails    = [r for r in results if r["emails"]]
    all_emails_set = {e for r in with_emails for e in r["emails"]}

    def _sort_date(r):
        try:
            return datetime.strptime(r["close_date"], "%d-%m-%Y")
        except Exception:
            return datetime.max

    # All opportunities sorted oldest-close-date first
    rows_html = ""
    for r in sorted(results, key=_sort_date):
        if r["emails"]:
            email_cells = []
            for email in r["emails"]:
                if email.lower() in new_emails:
                    email_cells.append(
                        f"<span style='background:#ffadb5;color:#8b0030;font-weight:bold;"
                        f"padding:1px 4px;border-radius:3px'>&#9733; {email}</span>"
                    )
                else:
                    email_cells.append(f"<a href='mailto:{email}'>{email}</a>")
            emails_html = ", ".join(email_cells)
            row_bg = "#fff3f4" if any(e.lower() in new_emails for e in r["emails"]) else ""
        else:
            emails_html = ""
            row_bg = ""

        rows_html += (
            f"<tr style='background:{row_bg}'>"
            f"<td style='padding:4px 8px;border:1px solid #d0e0dc'>"
            f"<a href='{r['url']}'>{r['number']}</a></td>"
            f"<td style='padding:4px 8px;border:1px solid #d0e0dc'>{r['buyer']}</td>"
            f"<td style='padding:4px 8px;border:1px solid #d0e0dc'>{r['title']}</td>"
            f"<td style='padding:4px 8px;border:1px solid #d0e0dc;white-space:nowrap'>{r['close_date']}</td>"
            f"<td style='padding:4px 8px;border:1px solid #d0e0dc'>{emails_html}</td>"
            f"</tr>\n"
        )

    new_badge = ""
    if new_emails:
        new_badge = (
            f"&nbsp;|&nbsp;<strong style='color:#8b0030'>&#9733; {len(new_emails)} new "
            f"contact{'s' if len(new_emails) != 1 else ''}</strong>"
        )

    html = f"""
<html><body style="font-family:Arial,sans-serif;font-size:14px;color:#222">
<h2 style="color:#19473c">BuyICT Opportunities - {run_date}</h2>
<p style="margin:8px 0">
  <strong>{len(results)}</strong> opportunities scraped &nbsp;|&nbsp;
  <strong>{len(with_emails)}</strong> with contact emails &nbsp;|&nbsp;
  <strong>{len(all_emails_set)}</strong> unique addresses
  {new_badge}
</p>
{"<p style='background:#fff3f4;border-left:4px solid #ffadb5;padding:8px 12px;margin:8px 0'>"
 "<strong>&#9733; New contacts this run:</strong> " + ", ".join(sorted(new_emails)) + "</p>"
 if new_emails else ""}
<p style="margin:8px 0;font-size:13px">
  <a href="{CONFIG['pages_url']}" style="color:#1f735e">View live dashboard &rarr;</a>
</p>
<table style="border-collapse:collapse;width:100%">
<thead>
<tr style="background:#19473c;color:#fff">
  <th style="padding:6px 8px;text-align:left">ID</th>
  <th style="padding:6px 8px;text-align:left">Buyer</th>
  <th style="padding:6px 8px;text-align:left">Title</th>
  <th style="padding:6px 8px;text-align:left">Closes</th>
  <th style="padding:6px 8px;text-align:left">Contact Email</th>
</tr>
</thead>
<tbody>
{rows_html}
</tbody>
</table>
<hr/>
<p style="color:#888;font-size:12px">
  &#9733; = new contact address (added to registry today) &nbsp;|&nbsp;
  Registry: buyict_email_registry.xlsx (attached) &nbsp;|&nbsp;
  Generated {now_aest().strftime("%Y-%m-%d %H:%M")}
</p>
</body></html>
"""

    text_lines = [f"BuyICT Opportunity Contacts — {run_date}", ""]
    if new_emails:
        text_lines += ["NEW EMAILS THIS RUN:", *sorted(new_emails), ""]
    for r in sorted(results, key=_sort_date):
        marker = "NEW: " if any(e.lower() in new_emails for e in r["emails"]) else "     "
        emails_str = ", ".join(r["emails"]) if r["emails"] else ""
        text_lines.append(f"{marker}{r['close_date']} | {r['number']} | {r['buyer']} | {emails_str}")
    plain = "\n".join(text_lines)

    new_count_str = f", {len(new_emails)} new" if new_emails else ""
    msg = MIMEMultipart("mixed")
    msg["Subject"] = (
        f"BuyICT Opportunities - {run_date} ({len(with_emails)} opportunities{new_count_str})"
    )
    msg["From"] = f"Market Analysis Tool <{CONFIG['from_email']}>"
    msg["To"]   = CONFIG["to_email"]

    body = MIMEMultipart("alternative")
    body.attach(MIMEText(plain, "plain"))
    body.attach(MIMEText(html, "html"))
    msg.attach(body)

    registry_path = Path(CONFIG["registry_file"])
    if registry_path.exists():
        with open(registry_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{registry_path.name}"')
        msg.attach(part)

    log.info(f"Sending email to {CONFIG['to_email']}...")
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(CONFIG["from_email"], password)
            smtp.sendmail(CONFIG["from_email"], CONFIG["to_email"], msg.as_string())
    except smtplib.SMTPAuthenticationError:
        with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
            smtp.ehlo(); smtp.starttls(); smtp.ehlo()
            smtp.login(CONFIG["from_email"], password)
            smtp.sendmail(CONFIG["from_email"], CONFIG["to_email"], msg.as_string())
    log.info("Email sent successfully.")


# ── Main ──────────────────────────────────────────────────────────────────────

async def main():
    import os
    password = CONFIG["gmail_app_password"] or os.environ.get("GMAIL_APP_PASSWORD", "")
    if not password:
        log.error("No Gmail App Password found. Set GMAIL_APP_PASSWORD env var or edit CONFIG.")
        sys.exit(1)

    log.info("=" * 60)
    log.info(f"BuyICT scraper starting — {now_aest().strftime('%Y-%m-%d %H:%M')}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            )
        )
        listing_page = await context.new_page()
        items = await collect_all_opportunity_urls(listing_page)
        await listing_page.close()

        log.info(f"Scraping {len(items)} detail pages (concurrency={CONFIG['concurrency']})...")
        results = await scrape_all_details(context, items)
        await browser.close()

    with_emails = [r for r in results if r["emails"]]
    log.info(f"Scrape complete: {len(with_emails)}/{len(results)} opportunities have emails")

    new_emails, _ = update_registry(results)
    log.info(f"New emails this run: {len(new_emails)}")

    html = generate_html(results, new_emails)
    save_html(html)

    send_email(results, new_emails, password)
    log.info("Done.")


if __name__ == "__main__":
    asyncio.run(main())
